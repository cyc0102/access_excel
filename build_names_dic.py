import openpyxl
fn = 'names.xlsx'
names_dict = {}
wb = openpyxl.load_workbook(fn)
# print(type(wb))
# allSheets = wb.get_sheet_names()
# print ("All sheets= ", allSheets)
# ws = wb.get_active_sheet()
ws = wb.get_sheet_by_name('Sheet1')
print("目前工作表= ",ws.title)
# print("B2=", ws['B2'].value)
# print("A2=", ws['A2'].value)
for i in range(2,279) :
    names_dict[ws.cell(column=2,row=i).value] = ws.cell(column=1,row=i).value

print("names_dict=",names_dict)    
